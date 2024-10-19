# encoding:utf-8
import time

import requests
import json
import webbrowser
import configparser
import hashlib
from config import conf,load_config,save_config
import os
from common.log import logger
from pydub import AudioSegment
from common.singleton import singleton
import openpyxl

class iPadWx:
    '''
    获取步骤：
    第一步：获取授权码(vx )
    第二步：领取一个机器人(单账号只能领取一次)
    第三步：获取登录二维码（准备好第四步再调用）
    第四步：确定已登录（确保扫码后调用）
    第五步：将auth_account和auth_password 填入config.json 文件
    '''
    shared_wx_contact_list = {}
    function_called = False
    def __init__(self):
        config = conf()


        self.base_url = config.get('base_url')
        self.auth_account = config.get('auth_account')
        self.auth_password = config.get('auth_password')
        self.token = config.get('token')
        self.auth = config.get('auth')

        self.is_auth = config.get_user_data('is_auth')
        self.is_create =config.get_user_data('is_create')
        self.is_login = config.get_user_data('is_login')
        self.is_first_login = config.get_user_data('is_first_login')




        # Try to load auth info from config
        if (not self.token or not self.auth) and self.is_login==1:
            if self.is_login==1:
                #todo 已经登录过再掉线，需要二次登录
                #self.relogin()
                print("微信已登录，不用重复登录")
            else: #没有登录过，需要扫码登录
                ret,data = self.login()
                if ret:
                    self.token = data.get("token")
                    self.auth = data.get("auth")
                    config.__setitem__("token", self.token)
                    config.__setitem__("auth", self.auth)
                    save_config()
                    config.user_datas.__setitem__("is_login",1)
                else:
                    pass
        if not iPadWx.function_called:
            self.load_contact()
            self.save_contact_xls()
            iPadWx.function_called = True


    def load_contact(self):
        if os.path.exists(f"contact_ipad.json"):
            iPadWx.shared_wx_contact_list = json.load(open(f"contact_ipad.json",'r',encoding='utf-8'))
            logger.info(f"读取联系人!")

        else:
            room_list = self.get_room_list()
            if room_list['code']==0:
                for room_id in room_list['data']:
                    if not room_id.endswith('@chatroom'):
                        continue
                    room_info = self.get_room_info(room_id)
                    iPadWx.shared_wx_contact_list[room_id] = room_info['data']
                    if room_info['code']==0:
                        if room_id.endswith('@chatroom'):
                            time.sleep(2)
                            members_info = self.get_chatroom_memberlist(room_id)
                            if members_info['code']==0:

                                room_info['data']['chatRoomMembers'] = members_info['data']

            # contacts = self.get_contact_list(0, 0)
            # if contacts['code'] == 0:
            #     current_wx_contact_seq = contacts['data']['current_wx_contact_seq']
            #     current_chat_room_contact_seq = contacts['data']['current_chat_room_contact_seq']
            #     wxids=[]
            #     while contacts['data']['ids']:
            #         for contact in contacts['data']['ids']:
            #             if contact not in ['weixin', 'fmessage', 'medianote', 'floatbottle',"filehelper","facesun"] \
            #                     and not contact.startswith("gh_") and not contact.endswith(
            #                 "@chatroom") :
            #                 wxids.append(contact)
            #         time.sleep(1)
            #         contacts = self.get_contact_list(current_wx_contact_seq, current_chat_room_contact_seq)
            #         current_wx_contact_seq = contacts['data']['current_wx_contact_seq']
            #         current_chat_room_contact_seq = contacts['data']['current_chat_room_contact_seq']
            #
            #     for i in range(0,len(wxids)//10 + 1):
            #         wxids_temp = wxids[i * 10:i * 10 + 10]
            #         wx_ids = ",".join(wxids_temp)
            #         time.sleep(2)
            #         contact_info = self.get_contact_info(wx_ids)
            #         for contact in contact_info['data']:
            #             iPadWx.shared_wx_contact_list[contact['userName']] = contact
            #
            #
            #     # for wxid in wxids:
            #     #
            #     #     #iPadWx.get_contact_info(wxids)
            #     #     iPadWx.shared_wx_contact_list[wxid] = iPadWx.shared_wx_contact_list.get(wxid, {})
            #
            #
            #
            #
            #     #iPadWx.shared_wx_contact_list.update(contacts['data'])
            #     #self.save_contact()

            self.save_contact()


    def save_contact(self):

        json.dump(iPadWx.shared_wx_contact_list,open(f"contact_ipad.json",'w',encoding='utf-8'),ensure_ascii=False, indent=4)
        logger.info(f"保存联系人!")

    def save_contact_xls(self):
        if os.path.exists(f"群资料.xlsx"):
            logger.info(f"群资料已经导出!")
            #return
        # 将 JSON 数据转换为 Python 对象
        data = iPadWx.shared_wx_contact_list

        # 创建一个新的 Excel 工作簿
        wb = openpyxl.Workbook()
        ws = wb.active

        # 写入表头
        ws.cell(row=1, column=1, value="群 ID")
        ws.cell(row=1, column=2, value="群名称")
        ws.cell(row=1, column=3, value="群人数")
        # 写入数据
        row_num = 2
        for group_id, group_info in data.items():
            if not group_id:
                continue
            ws.cell(row=row_num, column=1, value=group_id)
            ws.cell(row=row_num, column=2, value=group_info["nickName"])
            #members = ', '.join(group_info["group_members"])
            if group_id.endswith('@chatroom'):
                ws.cell(row=row_num, column=3, value=group_info["memberCount"])
            row_num += 1
        # 保存 Excel 文件
        wb.save('群资料.xlsx')

    # don't forget to import aiohttp at the beginning of your file


    # async def call_api(self, method, endpoint, data=None):
    #     url = f"{self.base_url}/{endpoint}"
    #     json = data if method == "POST" else None
    #
    #     async with aiohttp.ClientSession() as session:
    #         if method == "GET":
    #             async with session.get(url, headers=self.headers, params=data) as resp:
    #                 status = resp.status
    #                 response = await resp.text()
    #                 return json.loads(response)
    #         elif method == "POST":
    #             async with session.post(url, headers=self.headers, json=json) as resp:
    #                 status = resp.status
    #                 response = await resp.text()
    #                 return json.loads(response)

    def get_rooms(self):
        groups = self.get_room_list()

        if groups['code'] == 0:
            for room_id in groups['data']:
                logger.info(f"查询{room_id}")
                room_info = self.get_room_info(room_id)
                iPadWx.shared_wx_contact_list[room_id] = room_info['data']
                members = self.get_chatroom_memberlist(room_id)
                iPadWx.shared_wx_contact_list[room_id]['chatRoomMembers'] = members['data']
                logger.info(f"群还未查询过{room_id},名称{room_info['data']['nickName']}")
            self.save_contact()
        return iPadWx.shared_wx_contact_list


    def call_api(self, method, endpoint, data=None):
        url = f"{self.base_url}/{endpoint}"
        self.headers = {
            "we-token": self.token,
            "we-auth": self.auth,
            "Content-Type": "application/json"
        }
        try:
            if method == "GET":
                response = requests.get(url, headers=self.headers, params=data,timeout=100)
            elif method == "POST":
                response = requests.post(url, headers=self.headers, json=data,timeout=100)
            else:
                return None
            if response.status_code == 200:
                #logger.debug(response.json())
                return response.json()
            else:
                return None
        except Exception as e:
            logger.info("api 发送失败{0}".format(e))




    def create_user(self, user_data):
        return self.call_api("POST", "user/create", user_data)

    def get_qrcode(self, province=None, city=None):
        params = {
            "province": province
        }

        if city:
            params["city"] = city

        return self.call_api("GET", "user/qrcode", params)



    def encrypt_password(self, password):
        return hashlib.sha256((self.auth_account + password).encode()).hexdigest()

    def login(self):
        url = f"{self.base_url}/common/login"
        #encrypted_password = self.auth_password # self.encrypt_password(password)
        payload = {
            "phone": self.auth_account,
            "password": self.auth_password
        }
        headers = {
            "Content-Type": "application/x-www-form-urlencoded;charset=UTF-8"

        }
        self.headers = {
            "we-token": self.token,
            "we-auth": self.auth,
            "Content-Type": "application/json"
        }
        response = requests.post(url, data=(payload),headers=headers )

        if response.status_code == 200:
            data = response.json().get("data")
            if data:
                self.token = data.get("token")
                self.auth = data.get("auth")
                self.headers["we-token"] = self.token
                self.headers["we-auth"] = self.auth
                print(response.json())
                return True,data
            else:
                print(response.json())
        return False,None

    def extract_info_for_next_request(self):
        return {
            "token": self.token,
            "auth": self.auth
        }


    def generate_qr_code_html(self, qr_code_base64, output_file):
        # 从 base64 数据中提取图片数据
        image_data = qr_code_base64

        # 创建 HTML 页面
        html_content = f'''
        <!DOCTYPE html>
        <html>
        <head>
            <title>QR Code Login</title>
        </head>
        <body>
            <h1>请使用微信扫描下方二维码确认登录</h1>
            <img src="{image_data}" alt="QR Code Image">
        </body>
        </html>
        '''

        # 将 HTML 内容保存到指定文件中
        with open(output_file, 'w') as file:
            file.write(html_content)

        #print(f"HTML 文件已生成，请查看 {output_file} 文件来展示二维码图片。")
        logger.info(f"HTML 文件已生成，请查看 {output_file} 文件来展示二维码图片。")
        # 在浏览器中打开生成的 HTML 文件
        webbrowser.open(output_file)
    def confirm_login(self):
        #response = self.call_api("POST", "user/check")
        #if response['code']==0:
        #    pass
        return self.call_api("POST", "user/check")
    def relogin(self):
        return self.call_api("POST", "user/again" )

    def set_callback_url(self, url):
        # 在这里可以处理设置回调窗口的逻辑，例如保存回调URL
        payload = {
            "url": url,
        }
        print(f"Callback URL set to: {url}")

        return self.call_api("POST", "user/callback",payload)
    def start_listen(self):
        response = self.call_api("POST", "user/push/start")
        print(response)
        return response
    def stop_listen(self):
        response = self.call_api("POST", "user/push/stop")
        print(response)
        return response
    def get_user_info(self):
        return self.call_api("GET", "user/info")
    def get_robot_info(self):
        return self.call_api("GET", "user/robot/info")
    def filter_msg(self):
        '''
        最大10个
        '''
        payload = {
            "type": [
                "7001",
                "7005",
                "8001",
                "8002",
                "8003",
                "8004",
                "8005",
                "8006",
                "8007",
                "8008",
                "9001",
                "9002",
                "9003",
                "9004",
                "9005",
                "9006",
                "9007",
                "9008",
                "7099",

            ]
        }
        return  self.call_api("POST", "user/type/update",data=payload)

    def group_listen(self,payload):
        # payload = {
        #     "group": [
        #         "49670972421@chatroom",
        #         "26516713149@chatroom",
        #     ]
        # }
        return self.call_api("POST", "user/group/update", data=payload)
    def send_txt_msg(self,to_id,text):
        #print(text)
        #return
        payload = {
            "to_id": to_id,
            "content": text
        }
        return self.call_api("POST", "open/text/send", data=payload)
    def send_message(self,to_id,text):
        payload = {
            "to_id": to_id,
            "content": text
        }
        return self.call_api("POST", "open/text/send", data=payload)

    def send_at_message(self, to_id, at_ids, content):
        url = "open/text/at/send"

        data = {
            "to_id": to_id,
            "at_id": at_ids,
            "content": content
        }

        return self.call_api('POST',url, data)
    def send_at_msg(self, to_id, at_ids,nickname: str, content):
        '''
        发送@信息到群  send @ message to chatroom
        :param roomid: 群号(以@chatroom结尾) groupchatid(end with@chatroom)
        :param wxid: 要@的人的wxid(新用户的wxid以wxid_开头 老用户他们可能修改过 现在改不了) wechatid(start with wxid_) of person you want to @
        :param nickname: 要@的人的昵称，可随意修改 nickname of person you want to @
        :param content: 要发送的文本 content to send
        :return:
        '''
        #print(content)
        #return None
        url = "open/text/at/send"

        data = {
            "to_id": to_id,
            "at_id": at_ids,
            "content": content
        }

        return self.call_api('POST',url, data)
    def send_atall_message(self, to_id, content):
        '''
        发送@信息到群  send @ message to chatroom

        :return:
        '''
        url = "open/text/at-all/send"

        data = {
            "to_id": to_id,
            "content": content
        }

        return self.call_api('POST',url, data)
    def send_image_url(self, to_id, url):
        endpoint = "open/image/send"

        data = {
            "to_id": to_id,
            "url": url
        }

        return self.call_api('POST',endpoint, data)

    def send_pic_msg(self, to_id, url):
        endpoint = "open/image/send"

        data = {
            "to_id": to_id,
            "url": url
        }

        return self.call_api('POST', endpoint, data)
    def send_mini_program(self, to_id, xml):
        endpoint = "open/app/send"

        data = {
            "to_id": to_id,
            "xml": xml
        }

        return self.call_api('POST', endpoint, data)
    def send_emoji(self, to_id, md5,len):
        endpoint = "open/emoji/send"

        data = {
            "to_id": to_id,
            "md5": md5,
            "len":len
        }

        return self.call_api('POST', endpoint, data)
    def send_card(self, to_id, title,url,des,pic_url):
        endpoint = "open/card/send"

        data = {
            "to_id": to_id,
            "title": title,
            "url":url,
            "des":des,
            "pic_url":pic_url
        }

        return self.call_api('POST', endpoint, data)
    def send_mp3(self, to_id, title,mp3_url,des):
        endpoint = "open/mp3/send"

        data = {
            "to_id": to_id,
            "title": title,
            "mp3_url":mp3_url,
            "des":des
        }

        return self.call_api('POST', endpoint, data)


    def forward_video(self, to_id, xml):

        data = {'to_id': to_id, 'xml': xml}
        return self.call_api('POST', 'open/video/forward', data)
    def forward_img(self, to_id, xml):

        data = {'to_id': to_id, 'xml': xml}
        return self.call_api('POST', 'open/image/forward', data)
    def forward_file(self, to_id, title,total_len,attach_id,cdn_attach_url,aes_key,file_key,md5,file_ext):
        data ={
            "to_id": to_id,
            "title": title,
            "total_len": total_len,
            "attach_id": attach_id,
            "cdn_attach_url": cdn_attach_url,
            "file_ext": file_ext,
            "aes_key": aes_key,
            "file_key": file_key,
            "md5": md5
            }
        return self.call_api('POST', 'open/file/forward', data)

    def send_refer_msg(self, to_id, uuid, refer_id, refer_name,content):
        endpoint = "open/text/refer/send"

        data = {
            "to_id": to_id,
            "uuid": uuid,
            "refer_id": refer_id,
            "refer_name": refer_name,
            "content":content
        }

        return self.call_api('POST', endpoint, data)
    def revoke_message(self, to_id, uuid, msg_id):
        endpoint = "open/message/revoke"

        data = {
            "to_id": to_id,
            "uuid": uuid,
            "msg_id": msg_id
        }

        return self.call_api('POST', endpoint, data)
    def get_room_list2(self):
        return self.call_api('GET', 'open/room/list')

    def get_room_list(self):
        return self.call_api('GET', 'open/room/list')

    def get_room_info(self, room_id):
        params = {'room_id': room_id}
        return self.call_api('GET', 'open/room/info', params)

    def get_room_announcement(self, room_id):
        params = {'room_id': room_id}
        return self.call_api('GET', 'open/room/announcement', data=params)

    def get_room_member(self, room_id):
        params = {'room_id': room_id}
        return self.call_api('GET', 'open/room/member', data=params)
    def get_chatroom_memberlist(self, room_id):
        '''
        获取群成员后，保存到文件，方便下次获取
        :param room_id:
        :return:
        '''
        params = {'room_id': room_id}
        members = self.call_api('GET', 'open/room/member', data=params)

        #iPadWx.shared_wx_contact_list[room_id]['chatRoomMembers'] = members['data']
        #self.save_contact()
        return members

    def get_user(self,users, username):
        # 使用 filter 函数通过给定的 userName 来找寻符合条件的元素
        res = list(filter(lambda user: user['userName'] == username, users))

        return res[0] if res else None  # 如果找到了就返回找到的元素（因为 filter 返回的是列表，所以我们取第一个元素），否则返回 None


    def get_chatroom_nickname(self, room_id: str = 'null', wxid: str = 'ROOT'):
        '''
        获取群聊中用户昵称 Get chatroom's user's nickname
        群成员如果变了，没有获取到，则重新获取
        :param room_id: 群号(以@chatroom结尾) groupchatid(end with@chatroom)
        :param wxid: wxid(新用户的wxid以wxid_开头 老用户他们可能修改过 现在改不了) wechatid(start with wxid_)
        :return: Dictionary
        '''
        if room_id.endswith("@chatroom") and not wxid.endswith("@chatroom"):
            if room_id in iPadWx.shared_wx_contact_list:
                logger.debug("无需网络获取，本地读取")
                # logger.info(iPadWx.shared_wx_contact_list[room_id])
                member = iPadWx.shared_wx_contact_list[room_id]['chatRoomMembers']
                # logger.info(member)
                member_info = self.get_user(member, wxid)
                return member_info['displayName'], member_info['nickName']
        return None, None
    def invite_room_direct(self, room_id,wx_ids):
        data = {'room_id': room_id,
                "wx_ids": wx_ids
                }
        return self.call_api('POST', 'open/room/add', data=data)
    def invite_room_link(self, room_id,wx_ids):
        data = {'room_id': room_id,
                "wx_ids": wx_ids
                }
        return self.call_api('POST', 'open/room/invite', data=data)

    def create_room(self, room_name, wx_ids):
        data = {'room_name': room_name,
                "wx_ids": wx_ids
                }
        return self.call_api('POST', 'open/room/create', data=data)

    def exit_room(self, room_id):
        data = {'room_id': room_id}
        return self.call_api('POST', 'open/room/exit', data=data)

    def remove_member(self, room_id, wx_ids):
        data = {
            'room_id': room_id,
            'wx_ids': wx_ids
        }
        return self.call_api('POST', 'open/room/remove', data=data)

    def moment_friends(self, content, urls):
        data = {
            'content': content,
            'url': urls
        }
        return self.call_api('POST', 'open/circle/send', data=data)

    def get_contact_list(self, seq1,seq2):
        params = {'current_wx_contact_seq': seq1,
                  "current_chat_room_contact_seq":seq2
                  }
        return self.call_api('GET', 'open/user/all/list', params)
    def get_contact_info(self, wx_ids:str):
        params = {'wx_ids': wx_ids,
                  }
        return self.call_api('GET', 'open/user/info', params)
    def contact_tag_list(self):
        params = {
                  }
        return self.call_api('GET', 'open/tag/list', params)
    def contact_tag_update(self,tag_id,wx_ids):
        data = {
            "tag_id":tag_id,
            "wx_ids": wx_ids
                  }
        return self.call_api('POST', 'open/tag/update', data)
    def contact_tag_delete(self,tag_id):
        data = {
            "tag_id":tag_id
                  }
        return self.call_api('POST', 'open/tag/delete', data)
    def contact_tag_create(self,tag_name):
        data = {"tag_name":tag_name
                  }
        return self.call_api('POST', 'open/tag/create', data)
    def contact_operate(self,wx_id,type):
        '''
        {
            "wx_id": "wxid_id123123",
            "type": "4", //1 通讯录
            4 黑名单
            7 标星
            9 他是否可以看我的朋友圈
            10 消息免打扰
            12 置顶聊天
            17 不看他的朋友
            "switch": "on" //on 开启/拉黑等 off 关闭/取消等
            }
            '''
        data = {"wx_id":wx_id,
                "type":type
                }
        return self.call_api('POST', 'open/user/operate', data)

    def agree_friend(self, scene, ticket,wx_id):

        data = {
            "scene": scene,  # xml中获取
            "ticket": ticket,
            # xml中获取  朋友来源：2 来源邮箱; 3 来源微信号; 12 来源QQ; 13 来源通讯录 ;14 群聊; 15 手机号 ;17名片; 18 附近的人;25漂流瓶;29摇一摇;30二维码
            "wx_id": wx_id  # xml中获取
        }
        return self.call_api('POST', 'open/user/add', data=data)
    def upload_pic(self,local_file,upload_url):
        '''
        '{"url":"https://openai-75050.gzc.vod.tencent-cloud.com/openaiassets_14eb211d797ccdf33edc19839a7bcbcc_2579861717036942746.jpg","filekey":"openaiassets_14eb211d797ccdf33edc19839a7bcbcc_2579861717036942746.jpg"}'
        :param local_file:
        :param upload_url:
        :return:
        '''
        if os.path.exists(local_file):
            upload_url = "https://openai.weixin.qq.com/weixinh5/webapp/h774yvzC2xlB4bIgGfX2stc4kvC85J/cos/upload"
            payload = {}
            files = [
                    ('media', (
                    '0bbe70c70de11d13ffb2c.jpg', open(local_file, 'rb'),
                    'image/jpeg'))
                    ]
            headers={
            'Cookie':'oin:sess=eyJ1aWQiOiJmT1FMdFd1bUc0UyIsInNpZ25ldGltZSI6MTcxNjk2NjAwNTUxMCwiX2V4cGlyZSI6MTcxNzA1MjQwNjI4NiwiX21heEFnZSI6ODY0MDAwMDB9;oin:sess.sig=rHMmxTzNuo3bYpAwWzSoMA3S4DQ;wxuin=16966005491375;wxuin.sig=DMJDEtL7jcUjxUMAcDjetb9HBrA'
            }

            response=requests.post(upload_url,headers = headers,data = payload,files = files)

            print(response.text)
            return response.json()
        else:
            logger.info(f"本地文件不存在")
            return ""


